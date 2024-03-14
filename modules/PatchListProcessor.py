#import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
#import shutil
from datetime import datetime
import os
import re
# Function specifically for MSRC CVE table - To filter with colors
def process_patchlist(filename, importList):
    ProductList = importList # get the "must have" list
    try:
        wb = load_workbook(filename) 
        sheet = wb['Security Updates'] 

        # read the CVE List generated previously
        CVEList = []
        with open("CVE List.txt", "r") as txt_file:
            CVEList.extend([line.strip() for line in txt_file])
                # tqdm(temp, desc="Progress: ")
        for row in range(2, sheet.max_row + 1): 
            Product = sheet.cell(row, 2).value
            CVE_code = sheet.cell(row, 9).value
            if (CVE_code in CVEList):
                if (Product in ProductList) or (("Microsoft .NET Framework" in Product) and (sheet.cell(row, 3).value == "Windows Server 2016")): 
                    # situation 1: The software is in the "must-have" list
                    # situation 2: The software is .NET-related
                    sheet.cell(row, 2).fill = PatternFill("solid", start_color="FFa500")
                else: # 2nd priority: related to the current batch, but may not be required
                    sheet.cell(row, 2).fill = PatternFill("solid", start_color="c6efce")
            else: # last priority: unrelated to current batch, not usable/ no need to record
                 sheet.cell(row, 9).fill = PatternFill("solid", start_color="dcdcdc")
            # personal change: replace the word for better reading
            if (" (Server Core installation)" in Product):
                sheet.cell(row, 2).value = Product.replace("","")
            if ("-based Systems" in Product):
                sheet.cell(row, 2).value = re.sub(r' for .* Systems','',Product)

        wb.save(filename)
        # rename to Processed_Patch_List_YYYY-MM-DD
        os.rename(filename, str("Processed_Patch_List_{}.xlsx".format(datetime.now().date())))
        print ("Workbook handling completed.")
    except FileNotFoundError:
        print (" - Workbook not found, probably due to not downloading successfully.")
    
if __name__=="__main__": 
    print("Patch List Processor Called")